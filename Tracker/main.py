import numbers
import os
from http import server
import openpyxl, smtplib, ssl
from openpyxl import workbook, load_workbook
from email.message import EmailMessage
import tkinter as tk
from tkinter import messagebox, DISABLED
from openpyxl.cell import cell
from tkinter import *

def create_excel_file():
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet['A1'] = "Student Mail"
    sheet['B1'] = "Date"
    sheet['C1'] = "CI"
    sheet['D1'] = "Python"
    sheet['E1'] = "DM"
    workbook.save("C:\\Users\\cfryl\\PycharmProjects\\Tracker\\attendance.xlsx")

create_excel_file()

def save_to_excel(mail,date,Ci,Py,Dm):
    workbook = openpyxl.load_workbook("C:\\Users\\cfryl\\PycharmProjects\\Tracker\\attendance.xlsx")
    sheet = workbook.active
    sheet.append([mail, date, int(Ci), int(Py), int(Dm)])
    workbook.save("C:\\Users\\cfryl\\PycharmProjects\\Tracker\\attendance.xlsx")

def send_email(subject,body,to_email,sender_email,sender_password):
    message = EmailMessage()
    message.set_content(body)
    message['Subject'] = subject
    message['From'] = sender_email
    message['To'] = to_email

    context = ssl.create_default_context()

    with smtplib.SMTP("smtp.gmail.com", 587) as server:
        server.starttls(context=context)
        server.login(sender_email,sender_password)
        server.send_message(message)

def submit(mail_entry, date_entry, ci_entry, py_entry, dm_entry,sender_email,sender_password):
    mail = mail_entry.get()
    date = date_entry.get()
    try:
        ci = int(ci_entry.get())
        python = int(py_entry.get())
        dm = int(dm_entry.get())
    except ValueError:
        tk.messagebox.showwarning("Dikkat","Devamsızlıklar sayı olmalıdır.")
        return

    current_ci,current_python,current_dm = check_absences(mail)
    ci += current_ci
    python += current_python
    dm += current_dm

    save_to_excel(mail,date,ci,python,dm)       #birçok kişinin ismini oluşturmak için bu yerde


    if ci >= 2:
        if ci == 2:
            send_email("Dikkat", "CI dersi için sadece 1 devamsızlık hakkınız kaldı",mail,sender_email,sender_password)
        elif ci == 3:
            send_email("Dikkat", "CI dersi için devamsızlık hakkınız kalmadı",mail,sender_email,sender_password)
        elif ci > 3:
            send_email("Dikkat","CI dersi için devamsızlıktan kaldınız. Derse girme hakkınız yoktur.",mail,sender_email,sender_password)

    if python >= 1:
        if python == 1:
            send_email("Dikkat", "Python dersi için sadece 1 devamsızlık hakkınız kaldı",mail,sender_email,sender_password)
        elif python == 2:
            send_email("Dikkat", "Python dersi için devamsızlık hakkınız kalmadı",mail,sender_email,sender_password)
        elif python > 2:
            send_email("Dikkat","Python dersi için devamsızlıktan kaldınız. Derse girme hakkınız yoktur.",mail,sender_email,sender_password)

    if dm >= 1:
        if dm == 1:
            send_email("Dikkat", "DM dersi için sadece 1 devamsızlık hakkınız kaldı",mail,sender_email,sender_password)
        elif dm == 2:
            send_email("Dikkat", "DM dersi için devamsızlık hakkınız kalmadı",mail,sender_email,sender_password)
        elif dm > 2:
            send_email("Dikkat","DM dersi için devamsızlıktan kaldınız. Derse girme hakkınız yoktur.",mail,sender_email,sender_password)


    mail_entry.delete(0, tk.END)
    date_entry.delete(0, tk.END)
    ci_entry.delete(0, tk.END)
    py_entry.delete(0, tk.END)
    dm_entry.delete(0, tk.END)

    mail_entry.insert(0, "@gmail.com")
    mail_entry.config(fg='grey')
    date_entry.insert(0, "00/00/0000")
    date_entry.config(fg='grey')


def check_absences(mail):
    workbook = openpyxl.load_workbook("C:\\Users\\cfryl\\PycharmProjects\\Tracker\\attendance.xlsx")
    sheet = workbook.active

    ci=0
    python=0
    dm=0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] == mail:
            ci += int(row[2] or 0)
            python += int(row[3] or 0)
            dm += int(row[4] or 0)

    return ci, python, dm

def on_focus_in(event, entry, placeholder):
    if entry.get() == placeholder:
        entry.delete(0, tk.END)
        entry.config(fg='black')

def on_focus_out(event, entry, placeholder):
    if entry.get() == "":
        entry.insert(0, placeholder)
        entry.config(fg='grey')

def second_page(root,staff_email,staff_password):
    root.destroy()

    second_window = tk.Tk()
    second_window.title("Attendance Tracker")
    second_window.geometry("400x300")

    tk.Label(second_window, text="Student Mail").pack()
    mail_entry = tk.Entry(second_window, fg='grey')
    mail_entry.pack()
    mail_entry.insert(0,"@gmail.com")

    tk.Label(second_window, text="Date").pack()
    date_entry = tk.Entry(second_window, fg='grey')
    date_entry.pack()
    date_entry.insert(0, "00/00/0000")

    tk.Label(second_window, text="CI").pack()
    ci_entry = tk.Entry(second_window)
    ci_entry.pack()

    tk.Label(second_window, text="Python").pack()
    py_entry = tk.Entry(second_window)
    py_entry.pack()

    tk.Label(second_window, text="DM").pack()
    dm_entry = tk.Entry(second_window)
    dm_entry.pack()


    tk.Button(second_window, text="Submit", command=lambda: submit(mail_entry, date_entry, ci_entry, py_entry, dm_entry,staff_email,staff_password)).pack()

    mail_entry.bind('<FocusIn>', lambda event: on_focus_in(event, mail_entry, "@gmail.com"))
    mail_entry.bind('<FocusOut>', lambda event: on_focus_out(event, mail_entry, "@gmail.com"))
    date_entry.bind('<FocusIn>', lambda event: on_focus_in(event, date_entry, "00/00/0000"))
    date_entry.bind('<FocusOut>', lambda event: on_focus_out(event, date_entry, "00/00/0000"))



def main():
    root = tk.Tk()
    root.title("Attendance Tracker")
    root.geometry("400x300")


    tk.Label(root, text="Staff mail").pack()
    staff_email_entry = tk.Entry(root)
    staff_email_entry.pack()

    tk.Label(root, text="Password").pack()
    email_password_entry = tk.Entry(root,show="*")
    email_password_entry.pack()

    tk.Button(root, text="Log in", command=lambda: second_page(root,staff_email_entry.get(),email_password_entry.get())).pack()


    root.mainloop()

if __name__ == "__main__":
    main()






























